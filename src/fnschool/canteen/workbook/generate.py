import calendar
import io
import re
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import DecimalField, ExpressionWrapper, F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils.encoding import escape_uri_path
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fnschool import _, count_chinese_characters

from ..forms import (
    CategoryForm,
    ConsumptionForm,
    IngredientForm,
    PurchasedIngredientsWorkBookForm,
)
from ..models import Category, Consumption, Ingredient


def get_CNY_TEXT(amount):
    units = {
        "0": "\u96f6",  # ling2
        "1": "\u58f9",  # yi1
        "2": "\u8d30",  # er4
        "3": "\u53c1",  # san1
        "4": "\u8086",  # si4
        "5": "\u4f0d",  # wu3
        "6": "\u9646",  # liu4
        "7": "\u67d2",  # qi1
        "8": "\u634c",  # ba1
        "9": "\u7396",  # jiu3
    }

    levels = [
        "",
        "\u62fe",  # shi2
        "\u4f70",  # bai3
        "\u4edf",  # qian1
        "\u4e07",  # wan4
        "\u4ebf",  # yi4
        "\u5143",  # yuan2
        "\u89d2",  # jiao3
        "\u5206",  # fen1
        "\u6574",  # zheng3
    ]

    is_negative = False
    if amount < 0:
        is_negative = True
        amount = abs(amount)
    if amount == 0:
        return "\u96f6\u5143\u6574"  # ling2 yuan2 zheng3.

    amount = Decimal(str(amount)).quantize(
        Decimal("0.00"), rounding=ROUND_HALF_UP
    )
    amount_str = str(amount)

    integer_part = None
    decimal_part = None
    if "." in amount_str:
        integer_part, decimal_part = amount_str.split(".")
    else:
        integer_part = amount_str
        decimal_part = "00"

    result = []
    integer_part = integer_part.zfill(16)

    groups = [
        integer_part[-16:-12],
        integer_part[-12:-8],
        integer_part[-8:-4],
        integer_part[-4:],
    ]

    group_names = [
        "\u4e07",  # wan4
        "\u4ebf",  # yi4
        "\u4e07",  # wan4
        "\u5143",  # yuan2
    ]

    for i, group in enumerate(groups):
        group = group.lstrip("0")
        if not group:
            continue

        for j, digit in enumerate(group):
            if digit == "0":
                if result and result[-1] != "\u96f6":  # \\u96f6 is ling2 .
                    result.append("\u96f6")  # \\u96f6 is ling2 .
            else:
                result.append(units[digit])

                if len(group) - j - 1 > 0:
                    result.append(levels[len(group) - j - 1])

        if group_names[i]:
            result.append(group_names[i])

    if decimal_part != "00":
        if decimal_part[0] != "0":
            result.append(units[decimal_part[0]])
            result.append("\u89d2")  # \\u89d2 is jiao3 .

        if decimal_part[1] != "0":
            result.append(units[decimal_part[1]])
            result.append("\u5206")  # \\u5206 is fen1 .
    else:
        result.append("\u6574")  # \\u6574 is zheng3 .

    output = "".join(result)

    output = re.sub("\u96f6+", "\u96f6", output)
    output = re.sub("\u96f6([\u4e07\u4ebf])", r"\1", output)
    output = re.sub("\u96f6\u5143", "\u5143", output)
    output = re.sub("\u96f6\u89d2\u96f6\u5206", "", output)
    output = re.sub("\u96f6\u5206", "", output)

    if output.startswith("\u58f9\u62fe"):
        output = output.replace("\u58f9\u62fe", "\u62fe", 1)

    if is_negative:
        output = "\u8d1f" + output

    return output


def get_system_language():

    env_vars = ["LANG", "LC_ALL", "LC_MESSAGES", "LANGUAGE"]
    for var in env_vars:
        lang = os.environ.get(var)
        if lang:
            if "." in lang:
                lang = lang.split(".")[0]
            return lang
    return None


def is_zh_CN():
    lang = get_system_language()
    return lang == "zh_CN"


def set_column_width_in_inches(worksheet, column, inches):
    char_width = inches * 96 / 7

    if isinstance(column, int):
        col_letter = get_column_letter(column)
    else:
        col_letter = column
    worksheet.column_dimensions[col_letter].width = char_width


def set_row_height_in_inches(worksheet, row, inches):
    points = inches * 72
    worksheet.row_dimensions[row].height = points


class CanteenWorkBook:
    def __init__(self, request, month):
        self.wb = Workbook()
        self.wb[self.wb.sheetnames[0]].sheet_state = "hidden"
        self.cover_sheet = self.wb.create_sheet(title=_("Sheet Cover"))
        self.storage_sheet = self.wb.create_sheet(title=_("Sheet Storage"))
        self.storage_list_sheet = self.wb.create_sheet(
            title=_("Sheet Storage List")
        )
        self.non_storage_sheet = self.wb.create_sheet(
            title=_("Sheet Non-Storage")
        )
        self.non_storage_list_sheet = self.wb.create_sheet(
            title=_("Sheet Non-Storage List")
        )
        self.consumption_sheet = self.wb.create_sheet(
            title=_("Sheet Consumption")
        )
        self.consumption_list_sheet = self.wb.create_sheet(
            title=_("Sheet Consumption List")
        )
        self.surplus_sheet = self.wb.create_sheet(title=_("Sheet Surplus"))
        self.center_alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.font_16_bold = Font(size=16, bold=True)
        self.font_12 = Font(size=12)
        self.font_12_bold = Font(size=12, bold=True)

        self.request = request
        self.user = self.request.user
        self.year = int(month.split("-")[0])
        self.month = int(month.split("-")[1])
        self.date_start = datetime(self.year, self.month, 1).date()
        self.date_end = datetime(
            self.year, self.month, calendar.monthrange(self.year, self.month)[1]
        ).date()
        self.is_zh_CN = is_zh_CN()

        self.is_school = (
            (
                any(
                    [
                        name in self.user.affiliation
                        for name in [
                            "\u5e7c\u513f\u56ed",
                            "\u5c0f\u5b66",
                            "\u4e2d\u5b66",
                            "\u5927\u5b66",
                        ]
                    ]
                )
            )
            if is_zh_CN
            else False
        )

    def fill_in_storage_sheet(self):
        sheet = self.storage_sheet
        user = self.user
        title_cell = sheet.cell(1, 1)
        title_cell.value = _(
            "Table of {superior_department} Canteen Storaged Ingredients Statistics"
        ).format(
            superior_department=user.superior_department,
        )
        title_cell.font = self.font_16_bold
        title_cell.alignment = self.center_alignment
        for col_num, width in [
            [1, 2.23],
            [2, 3.14],
            [3, 3.14],
        ]:
            set_column_width_in_inches(sheet, col_num, width)
        sheet.merge_cells("A1:C1")

        sub_title_cell = sheet.cell(2, 1)
        sheet.merge_cells("A2:C2")
        sub_title_cell.font = self.font_12
        sub_title_cell.alignment = self.center_alignment
        sub_title_cell.value = _(
            "Affiliation: {affiliation}        Monetary Unit:         {year}.{month:0>2}.{day:0>2}"
        ).format(
            affiliation=user.affiliation,
            year=self.year,
            month=self.month,
            day=self.date_end.day,
        )

        header_row_num = 3
        header_category_cell = sheet.cell(header_row_num, 1)
        header_category_cell.value = _("Ingredient Categories (storage sheet)")

        header_total_price_cell = sheet.cell(header_row_num, 2)
        header_total_price_cell.value = _(
            "Ingredient Total Prices (storage sheet)"
        )

        header_note_cell = sheet.cell(header_row_num, 3)
        header_note_cell.value = _("Procurement Note")

        for cell in [
            header_category_cell,
            header_total_price_cell,
            header_note_cell,
        ]:
            cell.font = self.font_12
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()

        set_row_height_in_inches(sheet, 1, 0.38)
        set_row_height_in_inches(sheet, 2, 0.22)
        set_row_height_in_inches(sheet, 3, 0.32)

        for index, category in enumerate(categories):

            row_num = header_row_num + 1 + index
            set_row_height_in_inches(sheet, row_num, 0.32)

            category_cell = sheet.cell(row_num, 1)
            category_cell.value = category.name

            ingredients = Ingredient.objects.filter(
                Q(user=user)
                & Q(category=category)
                & Q(storage_date__gte=self.date_start)
                & Q(storage_date__lte=self.date_end)
                & Q(is_disabled=False)
                & Q(is_ignorable=False)
            ).all()
            total_price_cell = sheet.cell(row_num, 2)
            total_price_cell.value = sum([i.total_price for i in ingredients])

            note_cell = sheet.cell(row_num, 3)

            for cell in [category_cell, total_price_cell, note_cell]:
                cell.font = self.font_12
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(is_disabled=False)
            & Q(is_ignorable=False)
        ).all()

        summary_row_num = len(categories) + header_row_num + 1
        summary_total_price = sum([i.total_price for i in ingredients])
        summary_total_price_cell = sheet.cell(summary_row_num, 1)
        summary_total_price_cell.value = (
            _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=get_CNY_TEXT(summary_total_price),
                total_price=summary_total_price,
            )
            if is_zh_CN
            else _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=str(summary_total_price),
                total_price=summary_total_price,
            )
        )
        sheet.merge_cells(f"A{summary_row_num}:C{summary_row_num}")
        set_row_height_in_inches(sheet, summary_row_num, 0.32)

        handler_row_num = summary_row_num + 1
        handler_cell = sheet.cell(handler_row_num, 1)
        handler_cell.value = _("Handler:")
        sheet.merge_cells(f"A{handler_row_num}:C{handler_row_num}")
        set_row_height_in_inches(sheet, handler_row_num, 0.32)

        reviewer_row_num = handler_row_num + 1
        reviewer_cell = sheet.cell(reviewer_row_num, 1)
        reviewer_cell.value = _("Reviewer:")
        sheet.merge_cells(f"A{reviewer_row_num}:C{reviewer_row_num}")
        set_row_height_in_inches(sheet, reviewer_row_num, 0.32)

        supervisor_row_num = reviewer_row_num + 1
        supervisor_cell = sheet.cell(supervisor_row_num, 1)
        supervisor_cell.value = (
            _("Principal's Signature:")
            if self.is_school
            else _("Supervisor's Signature:")
        )
        sheet.merge_cells(f"A{supervisor_row_num}:C{supervisor_row_num}")
        set_row_height_in_inches(sheet, supervisor_row_num, 0.32)

        note_row_num = supervisor_row_num + 1
        note_cell = sheet.cell(note_row_num, 1)
        note_cell.value = (
            _(
                "Note: This form is a summary of all monthly food and "
                + "material inventory receipts from the cafeteria. After "
                + "verification, it will be signed and stamped with "
                + "the school seal by the principal as reimbursement "
                + "evidence."
            )
            if is_school
            else _(
                "Note: This form is a summary of all monthly food and "
                + "material inventory receipts from the cafeteria. "
                + "After verification, it will be signed and stamped "
                + "with the affiliation seal by the supervisor as "
                + "reimbursement evidence."
            )
        )
        sheet.merge_cells(f"A{note_row_num}:C{note_row_num}")
        set_row_height_in_inches(sheet, note_row_num, 0.27)

    def fill_in_cover_sheet(self):
        sheet = self.cover_sheet
        user = self.user
        title_cell = sheet.cell(1, 1)
        title_cell.value = _(
            "Table of {affiliation} Canteen Ingredients Procurement Statistics in {month:0>2} {year}"
        ).format(
            affiliation=user.affiliation,
            year=self.year,
            month=self.month,
        )
        title_cell.font = self.font_16_bold
        title_cell.alignment = self.center_alignment
        for col_num, width in [
            [1, 1.96],
            [2, 2.26],
            [3, 4.44],
        ]:
            set_column_width_in_inches(sheet, col_num, width)
        sheet.merge_cells("A1:C1")

        header_row_num = 2
        header_category_cell = sheet.cell(header_row_num, 1)
        header_category_cell.value = _("Ingredient Categories (cover sheet)")

        header_total_price_cell = sheet.cell(header_row_num, 2)
        header_total_price_cell.value = _("Ingredient Total Prices")

        header_note_cell = sheet.cell(header_row_num, 3)
        header_note_cell.value = _("Procurement Note")

        for cell in [
            header_category_cell,
            header_total_price_cell,
            header_note_cell,
        ]:
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            cell.font = self.font_12_bold

        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()

        set_row_height_in_inches(sheet, 1, 0.60)
        set_row_height_in_inches(sheet, 2, 0.44)

        for index, category in enumerate(categories):

            row_num = header_row_num + 1 + index
            set_row_height_in_inches(sheet, row_num, 0.44)

            category_cell = sheet.cell(row_num, 1)
            category_cell.value = category.name

            ingredients = Ingredient.objects.filter(
                Q(user=user)
                & Q(category=category)
                & Q(storage_date__gte=self.date_start)
                & Q(storage_date__lte=self.date_end)
                & Q(is_disabled=False)
            ).all()
            total_price_cell = sheet.cell(row_num, 2)
            total_price_cell.value = sum([i.total_price for i in ingredients])

            note_cell = sheet.cell(row_num, 3)
            note_cell.value = _(
                "Total price of storaged ingredients is {0}, total price of non-storaged ingredients is {1}."
            ).format(
                sum([i.total_price for i in ingredients if not i.is_ignorable]),
                sum([i.total_price for i in ingredients if i.is_ignorable]),
            )

            for cell in [category_cell, total_price_cell, note_cell]:
                cell.font = self.font_12
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(is_disabled=False)
        ).all()

        summary_row_num = len(categories) + header_row_num + 1
        summary_index_cell = sheet.cell(summary_row_num, 1)
        summary_index_cell.value = _("Procurement Summary")

        summary_total_price_cell = sheet.cell(summary_row_num, 2)
        summary_total_price_cell.value = sum(
            [i.total_price for i in ingredients]
        )

        summary_note_cell = sheet.cell(summary_row_num, 3)
        summary_note_cell.value = _(
            "Total price of storaged ingredients is {0}, total price of non-storaged ingredients is {1}."
        ).format(
            sum([i.total_price for i in ingredients if not i.is_ignorable]),
            sum([i.total_price for i in ingredients if i.is_ignorable]),
        )

        for cell in [
            summary_index_cell,
            summary_total_price_cell,
            summary_note_cell,
        ]:
            cell.font = self.font_12_bold
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        set_row_height_in_inches(sheet, summary_row_num, 0.44)

    def fill_in(self):
        self.fill_in_cover_sheet()
        # self.fill_in_storage_sheet()
        # self.fill_in_storage_list_sheet()
        # self.fill_in_non_storage_sheet()
        # self.fill_in_non_storage_list_sheet()
        # self.fill_in_consumption_sheet()
        # self.fill_in_consumption_list_sheet()
        # self.fill_in_surplus_sheet()
        # self.fill_in_food_sheets()

        return self.wb


def get_workbook(request, month):
    wb = CanteenWorkBook(request, month).fill_in()
    return wb
