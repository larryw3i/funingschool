import io
from datetime import datetime, date
import re
import pandas as pd
import numpy as np
from fnschool import count_chinese_characters
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from pathlib import Path
from fnschool import _
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render, redirect
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.encoding import escape_uri_path
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q

from .models import Ingredient
from .forms import PurchasedIngredientsWorkBookForm, IngredientForm

# Create your views here.

storage_date_header = (
    _("Storage Date"),
    _(
        'Formats like "YYYY.mm.dd", "YYYY/mm/dd", '
        + '"YYYY.mm.dd", "mm/dd", "mmdd", and "mm.dd" '
        + "are all acceptable. In short, FNSCHOOL wants to be "
        + "compatible with all the formats you like, but if "
        + "something goes wrong, you have to tell the "
        + "developers. Thank you!"
    ),
)
ingredient_name_header = (_("Ingredient Name"), _("Name of Ingredient"))
meal_type_header = (
    _("Meal Type"),
    _(
        "For example, breakfast, dinner, regular meals, "
        + "nutritious meals, etc., when generating a spreadsheet,"
        + " each meal category corresponds to a spread sheet. "
        + "If left blank, only one spreadsheet will be generated."
    ),
)

category_header = (
    _("Category"),
    _(
        "Usually there are seven categories: vegetables, meat, "
        + "grains, seasonings, eggs and milk, oils, and fruits."
    ),
)
quantity_header = (
    _("Quantity"),
    _(
        "When you purchase ingredients, it's best not to have "
        + "decimals in the quantity, as this can "
        + "be a big hassle!"
    ),
)


total_price_header = (
    _("Total Price"),
    None,
)
quantity_unit_name_header = (
    _("Unit Name of Quantity"),
    None,
)
is_ignorable_header = (
    _("Is Ignorable"),
    _("As long as a cell has content, it will be considered " + 'as "yes".'),
)


@login_required(login_url="/profiles/log_in")
def delete_ingredient(request, ingredient_id):
    ingredient = get_object_or_404(Ingredient, pk=ingredient_id)

    if request.method == "POST":
        if ingredient.user == request.user:
            ingredient.delete()
            return render(
                request,
                "close.html",
            )

    form = IngredientForm(instance=ingredient)
    return render(request, "canteen/delete_ingredient.html", {"form": form})


def edit_ingredient(request, ingredient_id):
    ingredient = get_object_or_404(Ingredient, pk=ingredient_id)

    if request.method == "POST":
        form = IngredientForm(request.POST, instance=ingredient)
        if form.is_valid():
            form.save()
            return render(
                request,
                "close.html",
            )
    else:
        form = IngredientForm(instance=ingredient)

    return render(request, "canteen/edit_ingredient.html", {"form": form})


date_patterns = [
    (r"\b\d{4}-\d{2}-\d{2}\b", "%Y-%m-%d"),
    (r"\b\d{4}/\d{2}/\d{2}\b", "%Y/%m/%d"),
    (r"\b\d{4}\.\d{2}\.\d{2}\b", "%Y.%m.%d"),
    (r"\b\d{8}\b", "%Y%m%d"),
]


@login_required
def list_ingredients(request):
    search_query = request.GET.get("q", "")
    search_query_cp = search_query
    fields = [
        f for f in Ingredient._meta.fields if f.name not in ["id", "user"]
    ]

    if search_query:
        queries = Q(user=request.user)

        search_query_dates = []

        for pattern, fmt in date_patterns:
            matches = re.findall(pattern, search_query)
            for match in matches:
                try:
                    date_obj = datetime.strptime(match, fmt).date()
                    search_query_dates.append(date_obj)
                    search_query = search_query.replace(match, "")

                except ValueError:
                    continue

        if len(search_query_dates) > 1:
            queries &= Q(storage_date__gte=min(search_query_dates))
            queries &= Q(storage_date__lte=max(search_query_dates))
        elif len(search_query_dates) == 1:
            queries &= Q(storage_date=search_query_dates[0])

        unit_names = Ingredient.objects.values("quantity_unit_name").distinct()
        unit_names = [
            c.get("quantity_unit_name")
            for c in unit_names
            if c.get("quantity_unit_name") in search_query
        ]
        for unit_name in unit_names:
            queries &= Q(quantity_unit_name__icontains=unit_name)
            search_query = search_query.replace(unit_name, "")

        categories = Ingredient.objects.values("category").distinct()
        categories = [
            c.get("category")
            for c in categories
            if c.get("category") in search_query
        ]
        for category in categories:
            queries &= Q(category__icontains=category)
            search_query = search_query.replace(category, "")

        meal_types = Ingredient.objects.values("meal_type").distinct()
        meal_types = [
            m.get("meal_type")
            for m in meal_types
            if m.get("meal_type") in search_query
        ]
        for meal_type in meal_types:
            queries &= Q(meal_type__icontains=meal_type)
            search_query = search_query.replace(meal_type, "")

        names = re.split(r"\s+", search_query)
        name_queries = Q()
        for name in names:
            name_queries |= Q(name__icontains=name)
        queries &= name_queries

        ingredients = Ingredient.objects.filter(queries)

    else:
        ingredients = Ingredient.objects.filter(Q(user=request.user))

    for f in fields:
        sort_name = request.GET.get("sort_" + f.name, "")
        if sort_name and sort_name in "+-":
            sort_name = (
                sort_name[1:] if sort_name.startswith("+") else sort_name
            )
            sort_name += f.name
            ingredients = ingredients.order_by(sort_name)

    per_page = request.GET.get("per_page")
    per_page = int(per_page) if str(per_page).isnumeric() else 10
    paginator = Paginator(ingredients, per_page)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)
    headers = [
        (f.name, request.GET.get("sort_" + f.name, ""), f.verbose_name)
        for f in fields
    ]
    context = {
        "page_obj": page_obj,
        "search_query": search_query_cp,
        "headers": headers,
        "per_page": per_page,
    }
    return render(request, "canteen/list_ingredients.html", context)


def create_ingredients(request):
    if request.method == "POST":
        form = PurchasedIngredientsWorkBookForm(request.POST, request.FILES)
        if form.is_valid():
            workbook_file = request.FILES["workbook_file"]

            if not workbook_file.name.endswith(".xlsx"):
                return HttpResponse(_('Please upload a file in "xlsx" format.'))

            df = pd.read_excel(workbook_file)

            for index, row in df.iterrows():
                Ingredient.objects.create(
                    user=request.user,
                    storage_date=row[storage_date_header[0]],
                    name=row[ingredient_name_header[0]],
                    meal_type=row[meal_type_header[0]],
                    category=row[category_header[0]],
                    quantity=row[quantity_header[0]],
                    total_price=row[total_price_header[0]],
                    quantity_unit_name=row[quantity_unit_name_header[0]],
                    is_ignorable=not row[is_ignorable_header[0]] == np.nan,
                )

            return redirect("canteen:list_ingredients")

    else:
        form = PurchasedIngredientsWorkBookForm()
    return render(request, "canteen/create_ingredients.html", {"form": form})


def get_template_workbook_of_purchased_ingredients(request):
    global storage_date_header, ingredient_name_header, meal_type_header
    global quantity_header, quantity_unit_name_header, total_price_header
    global is_ignorable_header
    headers = [
        storage_date_header,
        ingredient_name_header,
        meal_type_header,
        category_header,
        quantity_header,
        quantity_unit_name_header,
        total_price_header,
        is_ignorable_header,
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = _("Purchased Ingredients Sheet")

    for i, (h, c) in enumerate(headers):
        h_cell = ws.cell(1, i + 1)

        center_alignment = Alignment(horizontal="center", vertical="center")
        h_cell.alignment = center_alignment

        mono_font = Font(
            name="Mono",
            size=12,
        )
        h_cell.font = mono_font

        h_cell.value = h
        column_letter = get_column_letter(i + 1)
        hans_len = count_chinese_characters(h)
        hans_len = (hans_len + 2) if hans_len else hans_len
        ws.column_dimensions[column_letter].width = len(h) + hans_len + 2
        if c:
            h_cell.comment = Comment(c, _("the FNSCHOOL Authors"))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type=(
            "application/"
            + "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    today = datetime.now().date()
    last_month = today + relativedelta(months=-1, day=1)
    filename = (
        _("Purchased Ingredients WorkBook ({0})").format(
            last_month.strftime("%Y%m")
        )
        + ".xlsx"
    )

    encoded_filename = escape_uri_path(filename)
    response["Content-Disposition"] = (
        f'attachment; filename="{encoded_filename}"'
    )

    return response


# The end.
