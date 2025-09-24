import os
import sys
from openpyxl.utils.cell import get_column_letter
import tomllib
from tkinter import filedialog, ttk
import tkinter as tk

from fnschool import *
from fnschool.canteen.path import *
from fnschool.canteen.food import *
from fnschool.canteen.spreadsheet.base import *
from openpyxl.worksheet.datavalidation import DataValidation


class Purchasing(SsBase):
    def __init__(self, bill):
        super().__init__(bill)
        self.p_dpath_key = _("Parent Path Of Purchasing File")

    @property
    def wb(self):
        if not self._wb:
            print_info(_('Loading data from "{0}".').format(self.path))
            self._wb = load_workbook(self.path)
        return self._wb

    @wb.deleter
    def wb(self):
        self._wb = None
        self._sheet = None
        self._headers = None

    @property
    def sheet(self):
        if not self._sheet:
            self._sheet = self.wb[self.wb.sheetnames[0]]
        return self._sheet

    @property
    def food_classes(self):
        food_classes = self.bill.food_classes
        return food_classes

    def food_name_like(self, name, like):
        not_likes = None
        if "!" in like:
            like = like.split("!")
            not_likes = like[1:]
            like = like[0]

        result = None
        like_value = like.replace("*", "")
        if like.startswith("*") and not like.endswith("*"):
            result = name.endswith(like_value)
        elif like.endswith("*") and not like.startswith("*"):
            result = name.startswith(like_value)
        elif not "*" in like:
            result = like_value == name
        elif like.startswith("*") and like.endswith("*"):
            result = like_value in name

        if not_likes:
            result = result and not any(
                [self.food_name_like(name, nl) for nl in not_likes]
            )
        return result
        pass

    def get_food_class(self, name):
        food_classes = self.food_classes
        for fclass, name_likes in food_classes.items():
            for name_like in name_likes:
                if self.food_name_like(name, name_like):
                    return fclass
        return "蔬菜类"

    @property
    def path(self):
        p_dpath = self.cfg.get(self.p_path_key)
        initialdir = (
            p_dpath
            if (p_dpath and Path(p_dpath).exists())
            else ((Path.home() / "Downloads").as_posix())
        )
        if not self._path:
            self.gui.show_info(
                _(
                    "{0} need a purchasing list file, "
                    + "and it's file type should be '.xlsx'. "
                    + "The column names of it:"
                ).format(app_name)
                + _(
                    ""
                    + "\n\t ______________________________________________"
                    + "\n\t| column       | type  | example   | Note      |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Checking Date| Text  | 2024-03-01|           |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Food Name    | Text  | Cabbage   |           |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Meal Type    | Text  | Breakfast | Optional  |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Count        | Number| 20        |           |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Unit         | Text  | jin1      |           |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Total Price  | Number| 20.0      |           |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Is Negligible| Text  | y         |           |"
                    + "\n\t|``````````````|```````|```````````|```````````|"
                    + "\n\t| Is Inventory | Text  | y         |           |"
                    + "\n\t ``````````````````````````````````````````````"
                )
            )

            filetypes = ((_("Spreadsheet Files"), "*.xlsx"),)

            tkroot = tk.Tk()
            tkroot.withdraw()
            tkroot.attributes('-topmost', True)

            filename = filedialog.askopenfilename(
                parent=tkroot,
                title=_("Please select the purchasing file"),
                initialdir=initialdir,
                filetypes=filetypes,
            )

            if filename is None or filename == ():
                print_warning(_("No file was selected, exit."))
                exit()
                return None
            print_info(
                _('Purchasing list file "{0}" has been selected.').format(
                    filename
                )
            )
            self._path = filename
        self.cfg.set(self.p_dpath_key, Path(self._path).parent.as_posix())
        return self._path

    def update_data_validations(self):
        if not self.food_class_dv in self.sheet.data_validations:
            self.sheet.add_data_validation(self.food_class_dv)

    def update(self):
        self.update_data_validations()
        self.add_class_col()
        self.update_inventories()

    def get_meal_types(self):
        data = self.pd_data
        col = self.meal_type_col
        if not col:
            return []
        self._meal_types = list(set(data.loc[:, col[0]]))
        if len(self._meal_types) > 0:
            return self._meal_types
        return None

    def update_inventories(self):

        merged_ranges = list(self.sheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            sheet.unmerge_cells(str(cell_group))

        inventories_len = len(
            [
                row_index
                for row_index in range(2, self.sheet.max_row + 1)
                if self.sheet.cell(row_index, self.inventory_col[1]).value
            ]
        )

        if inventories_len > 0:
            return

        print_warning(
            _("The remaining food wasn't read " + 'from "{0}".').format(
                self.path,
            )
        )

        meal_types = self.get_meal_types()
        inventory = self.bill.spreadsheet.inventory

        update_yn = False

        for meal_type in meal_types:
            bill_fpath = self.operator.get_bill_fpath(meal_type)
            print_info(
                _(
                    "{0} is reading remaining foods from "
                    + 'Sheet "{1}" of Spreadsheet "{2}" ......'
                ).format(
                    app_name,
                    inventory.sheet_name,
                    bill_fpath,
                )
            )
            saved_ifoods = inventory.get_save_foods(meal_type)
            saved_ifoods_len = len(saved_ifoods)
            if saved_ifoods_len < 1:
                print_warning(
                    _('There is no saved inventories from "{0}"').format(
                        bill_fpath
                    )
                )
                return

            print_warning(
                (
                    _(
                        "Some remaining foods have been read "
                        + 'from sheet "{0}" of spreadsheet "{1}":'
                    )
                    if len(saved_ifoods) > 1
                    else _(
                        "The remaining food has been read"
                        + 'from sheet "{0}" of spreadsheet "{1}":'
                    )
                ).format(inventory.sheet_name, bill_fpath)
            )

            saved_ifoods_len2 = len(str(saved_ifoods_len))
            saved_ifoods_s = sqr_slist(
                [
                    (
                        f"({i+1:>{saved_ifoods_len2}}) "
                        + f"{f0.name}:{f0.count} {f0.unit_name}"
                        + f"\u2a09 {f0.unit_price:.2f} "
                        + f"{self.bill.currency.unit}/"
                        + f"{f0.unit_name}={f0.total_price:.2f} "
                        + f"{self.bill.currency.unit}"
                    )
                    for i, f0 in enumerate(saved_ifoods)
                ]
            )

            saved_ifoods_s_len = max(
                [len(s) for s in saved_ifoods_s.split("\n")]
            )
            saved_ifoods_info = (
                _("Purchaser: ")
                + saved_ifoods[0].purchaser
                + "\n"
                + _("Inventory data: ")
                + saved_ifoods[0].xdate.strftime("%Y.%m.%d")
                + "\n"
            )
            print_info(saved_ifoods_s)
            print_warning(saved_ifoods_info)
            print_warning(
                (
                    _('Fill them in "{0}"? (YyNn, default: "No")')
                    if len(saved_ifoods) > 1
                    else _('Fill it in "{0}"? (YyNn, default: "No")')
                ).format(self.path)
            )

            f_input = get_input().replace(" ", "")
            if len(f_input) > 0 and f_input in "Yy":
                update_yn = True
                max_row = len(
                    [
                        row_index
                        for row_index in range(1, self.sheet.max_row + 1)
                        if self.sheet.cell(row_index, 1).value
                    ]
                )
                for row_index in range(
                    max_row + 1, max_row + 1 + len(saved_ifoods)
                ):
                    f = saved_ifoods[row_index - max_row - 1]
                    self.sheet.cell(
                        row_index, self.xdate_col[1]
                    ).number_format = numbers.FORMAT_TEXT
                    self.sheet.cell(
                        row_index,
                        self.xdate_col[1],
                        f.xdate.strftime("%Y-%m-%d"),
                    )
                    self.sheet.cell(
                        row_index, self.purchaser_col[1], f.purchaser
                    )
                    self.sheet.cell(row_index, self.food_name_col[1], f.name)
                    self.sheet.cell(row_index, self.food_class_col[1], f.fclass)
                    self.food_class_dv.add(
                        self.sheet.cell(row_index, self.food_class_col[1])
                    )
                    self.sheet.cell(
                        row_index,
                        self.unit_name_col[1],
                        f.unit_name,
                    )
                    self.sheet.cell(
                        row_index, self.count_col[1]
                    ).number_format = numbers.FORMAT_NUMBER_00
                    self.sheet.cell(row_index, self.count_col[1], f.count)
                    self.sheet.cell(
                        row_index, self.total_price_col[1]
                    ).number_format = numbers.FORMAT_NUMBER_00
                    self.sheet.cell(
                        row_index,
                        self.total_price_col[1],
                        f.total_price,
                    )
                    self.sheet.cell(row_index, self.inventory_col[1], "y")

                    for col_index in self.col_indexes:
                        if not col_index:
                            continue
                        self.sheet.cell(row_index, col_index).font = (
                            self.edited_cell_font
                        )

                print_info(
                    (
                        _("The remaining foods have been " + 'added to "{0}".')
                        if len(saved_ifoods) > 1
                        else _('The remaining food has been added to "{0}".')
                    ).format(self.path)
                )

                pass

        self.wb.save(self.path)
        self.wb.close()
        del self.wb

        if update_yn:
            print_info(
                (
                    _(
                        "Please check/modify the updated data. "
                        + "(Press any key to open the file)"
                    )
                )
            )
            get_input()
            open_path(self.path)
            print_info(
                _("Ok, I checked it, it's ok. " + "(Press any key to continue)")
            )
            get_input()
            pass

        pass

    def split_foods(self):
        foods_cp = self.bill.foods.copy()
        foods = self.bill.foods

        bad_total_price_foods = []
        for f in foods:
            total_price_s = str(f.total_price)
            if "." in total_price_s and len(total_price_s.split(".")[1]) > 2:
                bad_total_price_foods.append(f)
        if len(bad_total_price_foods) > 0:
            bad_total_price_foods_len = len(bad_total_price_foods)
            print_error(
                (
                    _(
                        "The total price of the following food has more than "
                        + "two decimal places, and {0} cannot process it."
                    )
                    if bad_total_price_foods_len == 1
                    else _(
                        "The total price of the following foods have more than "
                        + "two decimal places, and {0} cannot process them."
                    )
                ).format(app_name)
            )
            bad_total_price_foods_len2 = len(str(bad_total_price_foods_len + 1))
            print_error(
                sqr_slist(
                    [
                        (
                            f"{i+1:>{bad_total_price_foods_len2}} "
                            + _("{0}({1})").format(
                                f.name,
                                (
                                    f"{f.xdate.year}.{f.xdate.month:0>2}"
                                    + f".{f.xdate.day:0>2}"
                                ),
                            )
                            + f" {f.total_price}"
                        )
                        for i, f in enumerate(bad_total_price_foods)
                    ]
                )
            )
            print_error(_("Exit."))
            exit()

        split_mode = ""

        for i, f in enumerate(foods_cp):
            up0, up1, threshold = f.count_threshold
            f_count = f.count
            f_total_price = f.total_price
            f_unit_price = f.unit_price
            if threshold:
                if split_mode in "YyNn":
                    print_info(
                        _(
                            'The unit price of "{0}" is an '
                            + 'infinite decimal, split "{0}"?'
                            + '(Yes: "Y","y","".'
                            + ' Yes for rest: "A","a".'
                            + ' No: "N","n".'
                            + ' No for rest: "S","s".'
                            + ' Default: Yes for rest, "A/a".'
                            + ")"
                        ).format(f.name)
                    )
                    split_mode = get_input()

                if split_mode and split_mode in "Ss":
                    return

                if split_mode == "":
                    split_mode = "A"

                if split_mode in "YyAa":
                    times_char = "\u2a09"
                    f0 = foods[i]
                    f0.count = threshold
                    f0.total_price = up1 * threshold
                    f0._count_threshold = None

                    f1_count = f_count - threshold
                    f1 = Food(
                        self.bill,
                        name=f0.name,
                        unit_name=f0.unit_name,
                        count=f1_count,
                        total_price=f1_count * up0,
                        xdate=f0.xdate,
                        purchaser=f0.purchaser,
                        fclass=f0.fclass,
                        is_inventory=f0.is_inventory,
                        is_abandoned=f0.is_abandoned,
                        meal_type=f0.meal_type,
                    )

                    total_price0 = round(
                        up1 * threshold + f1_count * up0, self.sd + 1
                    )
                    print_info(
                        _('"{0}" was split:').format(f0.name)
                        + f"\n\t{f0.unit_price} "
                        + f"{self.bill.currency.unit}"
                        + f"/{f0.unit_name} "
                        + f"{times_char} {f0.count} "
                        + f"{f0.unit_name} + "
                        + f"{f1.unit_price} "
                        + f"{self.bill.currency.unit}"
                        + f"/{f1.unit_name} "
                        + f"{times_char} {f1.count} "
                        + f"{f1.unit_name} = "
                        + f"{total_price0} "
                        + f"{self.bill.currency.unit} = "
                        + f"{f_unit_price} "
                        + f"{self.bill.currency.unit}"
                        + f"/{f.unit_name} "
                        + f"{times_char} {f_count} "
                        + f"{f.unit_name} = "
                        + f"{f_total_price}"
                        + f"{self.bill.currency.unit}"
                    )

                    foods.append(f1)
        return

    def abandoned_foods_pass(self, foods):
        ab_foods = [
            f
            for f in foods
            if f.is_abandoned and not f.fclass in self.condiment_class_names
        ]
        if len(ab_foods) < 1:
            return True
        print_warning(
            (
                _(
                    "Normally, {0} only considers non-warehoused foods as "
                    + "seasonings, but {0} has found that the following food "
                    + "is non-warehoused, but it is not seasonings:"
                )
                if len(ab_foods) == 1
                else _(
                    "Normally, {0} only considers non-warehoused foods as "
                    + "seasonings, but {0} has found that the following foods "
                    + "are non-warehoused, but they are not seasonings:"
                )
            ).format(app_name)
        )

        print_info(
            sqr_slist([_("{0}({1})").format(f.name, f.xdate) for f in ab_foods])
        )
        print_error(
            (
                _(
                    "You may need to modify its food class, "
                    + "otherwise {0} will go wrong."
                )
                if len(ab_foods) == 1
                else _(
                    "You may need to modify their food classes, "
                    + "otherwise {0} will go wrong."
                )
            ).format(app_name)
            + _(
                ' Do you want to re-edit "{0}"? '
                + "(Yes: 'Y' or 'y'. Default: No.)"
            ).format(self.path)
        )
        re_edit_yn = get_input()
        if re_edit_yn and re_edit_yn in "Yy":
            open_path(self.path)
            print_info(
                _(
                    "Ok, I have re-edited and closed it? "
                    + "(Press any key to continue)"
                )
            )
            get_input()
            return False

        return True

        pass

    def read_pfoods(self):
        self.update()
        foods = self.pd_data
        # if not self.abandoned_foods_pass(foods):
        #     self.pd_data = None
        #     return self.read_pfoods()
        _foods = []
        for __, food in foods.iterrows():
            _food = Food(
                self.bill,
                name=food[self.food_name_col[0]],
                unit_name=food[self.unit_name_col[0]],
                count=food[self.count_col[0]],
                total_price=food[self.total_price_col[0]],
                xdate=food[self.xdate_col[0]],
                purchaser=food[self.purchaser_col[0]],
                fclass=food[self.food_class_col[0]],
            )
            if self.abandoned_col[0]:
                _food.is_abandoned = not pd.isna(food[self.abandoned_col[0]])
            if self.inventory_col[0]:
                _food.is_inventory = not pd.isna(food[self.inventory_col[0]])
            if self.meal_type_col and self.meal_type_col[0]:
                _food.meal_type = food[self.meal_type_col[0]]

            _foods.append(_food)

        foods = _foods
        foods = sorted(foods, key=lambda f: f.xdate)
        self.bill.foods = foods

        self.split_foods()

        self.spreadsheet.preconsuming.pre_consume_foods()

        return foods
        pass


# The end.
