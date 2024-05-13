import os
import sys
from openpyxl.styles import *
from openpyxl.formatting.rule import *
from openpyxl.styles.differential import *
from openpyxl.utils.cell import *


from fnschool import *
from fnschool.canteen.spreadsheet.base import *


class Unwarehousing(SpreadsheetBase):
    def __init__(self, bill):
        super().__init__(bill)
        self.sheet_name = "未入库明细表"
        pass

    @property
    def form_indexes(self):
        if not self._form_indexes:
            unwsheet = self.sheet
            indexes = []
            row_index = 1
            for row in unwsheet.iter_rows(
                max_row=unwsheet.max_row + 1, max_col=7
            ):
                if row[0].value and "未入库明细表" in row[0].value.replace(
                    " ", ""
                ):
                    indexes.append([row_index + 1, 0])

                if row[1].value and "合计" in row[1].value.replace(" ", ""):
                    indexes[-1][1] = row_index

                row_index += 1

            if len(indexes) > 0:
                self._form_indexes = indexes
            else:
                return None

        return self._form_indexes

    def update(self):
        unwsheet = self.sheet
        form_indexes = self.form_indexes

        foods = [f for f in self.bfoods if f.is_abandoned]
        if len(foods) < 1:
            print_info(
                _(
                    "There is no abandoned food. "
                    + "Sheet {0} updating skipped."
                ).format(self.sheet.title)
            )
            return None

        foods = sorted(foods, key=lambda f: f.xdate)

        t1 = []

        for bf in self.bfoods:
            for d, __ in bf.consumptions:
                t1.append(d)
        t1 = sorted(t1)[-1]

        row_indexes = []
        for form_index in form_indexes:
            form_index0, form_index1 = form_index
            unwsheet.cell(form_index0, 1, f" 学校名称：{self.purchaser}")
            unwsheet.cell(
                form_index0,
                4,
                f"        "
                + f"{t1.year} 年 {t1.month} 月 "
                + f"{t1.day} 日"
                + f"               ",
            )
            row_index_start = form_index0 + 2
            row_index_end = form_index1 - 1
            row_indexes += list(range(row_index_start, row_index_end + 1))

        for row_index in row_indexes:
            for col_index in range(1, 7 + 1):
                unwsheet.cell(row_index, col_index, "")

        total_price = 0.0
        use_forms = False

        for _index, row_index in enumerate(row_indexes):
            food = foods[_index]
            total_price += food.total_price
            unwsheet.cell(row_index, 1, food.xdate.strftime("%Y.%m.%d"))
            unwsheet.cell(row_index, 2, food.name)
            unwsheet.cell(row_index, 3, food.unit_name)
            unwsheet.cell(row_index, 4, food.count)
            unwsheet.cell(row_index, 5, food.unit_price)
            unwsheet.cell(row_index, 6, food.total_price)
            unwsheet.cell(row_index, 5).number_format = numbers.FORMAT_NUMBER_00
            unwsheet.cell(row_index, 6).number_format = numbers.FORMAT_NUMBER_00
            if (
                str(unwsheet.cell(row_index + 1, 2).value)
                .replace(" ", "")
                .endswith("合计")
                and len(foods) - 1 > _index
            ):
                unwsheet.cell(row_index + 1, 2, "合计")
                unwsheet.cell(row_index + 1, 6, total_price)
                use_forms = True

            if len(foods) - 1 == _index:
                for row in unwsheet.iter_rows(
                    min_row=row_index,
                    max_row=unwsheet.max_row,
                    min_col=1,
                    max_col=7,
                ):
                    if row[2].value and str(row[2].value).replace(
                        " ", ""
                    ).endswith("合计"):
                        row[2].value = "总合计" if use_forms else "合计"
                        row[6].value = total_price
                        break
                break

        print_info(_("Sheet '%s' was updated.") % self.sheet.title)


# The end.
