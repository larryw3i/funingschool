import os
import sys

import pandas as pd
import numpy as np
import uuid

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.formatting.rule import *
from openpyxl.styles.differential import *
from openpyxl.utils.cell import *


from openpyxl.utils import range_boundaries
from fnschool import *
from fnschool.canteen import *
from fnschool.canteen.food import *
from fnschool.canteen.bill import *
from fnschool.canteen.path import *


class WorkBook:
    def __init__(self, bill):
        self.bill = bill
        self.check_sheet_name = "清点表"
        self.unit_sheet_name = "计量单位表"
        self.warehousing_sheet_name = "入库单"
        self.unwarehousing_sheet_name = "未入库明细表"
        self.consuming_sum_sheet_name = "出库汇总表"
        self.consuming_sheet_name = "出库单"
        self.inventory_sheet_name = "食材盘存表"
        self.food_sheet0_name = "材料台账母表"
        self.pre_consuming_sheet0_name = "简易出库母表"
        self.base_class_sheet_name = "大类表"
        self.pre_consuming_sheet_name_prefix = "出库表"
        self.purchase_sum_sheet_name = "入库、未入库汇总表"
        self.cover_sheet_name = "六大类总封面"
        self.purchase_sheet_names = ["客户商品销售报表", "客户送货明细报表"]
        self.warehousing_form_index_offset = 0
        self.inventory_form_index_offset = 1
        self._workbook = None
        self._main_spreadsheet_path = None
        self._check_df = None
        self._unit_name_list = None
        self._unit_df = None
        self._negligible_class_list = None
        self._base_class_df = None
        self.purchase_workbook_fd_path = None
        self.pre_consuming_sheet_col_index_offset = 6
        self.pre_consuming_sheet_row_index_offset = 3
        self.spreadsheet_ext_names = ["xlsx"]
        self.cell_alignment0 = Alignment(
            horizontal="center", vertical="center"
        )
        self.cell_side0 = Side(border_style="thin")
        self.cell_border0 = Border(
            top=self.cell_side0,
            left=self.cell_side0,
            right=self.cell_side0,
            bottom=self.cell_side0,
        )

    @property
    def profile(self):
        return self.bill.profile

    @property
    def food(self):
        return self.bill.food

    def get_conver_sheet(self):
        return self.get_sheet(self.cover_sheet_name)

    def get_purchase_sum_sheet(self):
        return self.get_sheet(self.purchase_sum_sheet_name)

    def get_consuming_sum_sheet(self):
        return self.get_sheet(self.consuming_sum_sheet_name)

    def get_days_of_pre_consuming_sheet(self, name):
        time_start, time_end = self.get_time_node_of_pre_consuming_sheet(name)
        return (time_end - time_start).days + 1

    def get_time_node_of_pre_consuming_sheet(self, name):
        pcsheet = self.get_sheet(name)
        time_start = pcsheet.cell(
            1, self.pre_consuming_sheet_col_index_offset
        ).value.split(".")
        time_start = datetime(
            int(time_start[0]), int(time_start[1]), int(time_start[2])
        )
        time_node = [time_start, None]
        for col_index in range(
            self.pre_consuming_sheet_col_index_offset, pcsheet.max_column
        ):
            cell_value = pcsheet.cell(1, col_index).value
            if not cell_value:
                time_end = pcsheet.cell(1, col_index - 1).value.split(".")
                time_end = datetime(
                    int(time_end[0]), int(time_end[1]), int(time_end[2])
                )
                time_node[1] = time_end
                return time_node

        return None

    def get_base_class_names(self, include_negligible_name=False):
        df = self.get_base_class_df()
        names = df.index.tolist()
        return names[:-1]

    def get_base_class_names_include_negligible_name(self):
        return self.get_base_class_names(include_negligible_name=True)

    def get_base_class_df(self):
        return self.get_base_class_df_from_spreadsheet()

    def get_base_class_df_from_spreadsheet(self, sheet_name="大类表"):
        if not self._base_class_df is None:
            return self._base_class_df

        base_class_df = pd.read_excel(
            self.bill.workbook.get_main_spreadsheet_path(),
            sheet_name=sheet_name,
        )
        self._base_class_df = base_class_df.T
        return self._base_class_df

    def get_negligible_class_list(self):
        if not self._negligible_class_list is None:
            return self._negligible_class_list
        self._negligible_class_list = (
            self.get_base_class_df().loc["非入库类"].to_list()
        )
        return self._negligible_class_list

    def get_unit_name_list(self):
        if not self._unit_name_list is None:
            return self._unit_name_list
        unit_df = self.get_unit_df()
        self._unit_name_list = unit_df["Name"].tolist()
        return self._unit_name_list

    def get_sheet_names(self):
        wb = self.get_workbook()
        return wb.sheetnames

    def includes_sheet(self, sheet):
        if isinstance(sheet, str):
            name = sheet
        else:
            name = sheet.title
        names = self.get_sheet_names()
        if name.endswith("*"):
            name = name.replace("*", "")
            return any(_name.startswith(name) for _name in names)
        return name in names

    def get_pre_consuming_sheet0(self):
        return self.get_sheet(self.pre_consuming_sheet0_name)

    def get_consuming_n_by_time_node(self, time_node):
        time_nodes = self.bill.get_time_nodes()
        time_nodes = [t for t in time_nodes if t[0].month == time_node.month]
        n_index = 0
        for time_start, time_end in time_nodes:
            time_range = [
                time_start + timedelta(days=i)
                for i in range((time_end - time_start).days + 1)
            ]
            if time_node in time_range:
                n_index += time_range.index(time_node) + 1
            else:
                n_index += len(time_range)
        return n_index

    def format_inventory_sheet(self):
        isheet = self.get_inventory_sheet()
        self.unmerge_cells_of_sheet(isheet)

        for row in isheet.iter_rows(
            min_row=1, max_row=isheet.max_row, min_col=1, max_col=9
        ):
            isheet.row_dimensions[row[0].row].height = 14.25

            if row[8].value and "原因" in str(row[8].value).replace(" ", ""):
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=9,
                    end_column=9,
                )

            if row[6].value and str(row[6].value).replace(" ", "") == "差额栏":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=7,
                    end_column=8,
                )

            if row[4].value and str(row[4].value).replace(" ", "") == "盘点栏":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=5,
                    end_column=6,
                )

            if row[2].value and str(row[2].value).replace(" ", "") == "账面栏":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=3,
                    end_column=4,
                )

            if row[0].value and row[0].value.replace(" ", "") == "食材名称":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=1,
                    end_column=1,
                )

            if row[0].value and (
                "备注" in row[0].value.replace(" ", "")
                or "审核人" in row[0].value.replace(" ", "")
            ):
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=9,
                )

            if row[0].value and row[0].value.replace(" ", "") == "食材盘存表":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=9,
                )
                isheet.row_dimensions[row[0].row].height = 22.5

                isheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=1,
                    end_column=9,
                )

    def update_inventory_sheet_by_time_node_m1(self):
        isheet = self.get_inventory_sheet()
        foods = self.food.get_foods_from_pre_consuming_sheet_m1()
        foods = [f for f in foods if f.get_remainder() > 0.0]
        form_indexes = self.get_inventory_form_indexes()
        time_nodes = self.bill.get_time_nodes()
        time_start, time_end = time_nodes[-1]
        form_indexes_n = len(time_nodes) + self.inventory_form_index_offset - 1
        form_index = form_indexes[form_indexes_n]
        form_index_start, form_index_end = form_index
        fentry_index_start = form_index_start + 3
        fentry_index_end = form_index_end - 1

        self.unmerge_cells_of_sheet(isheet)

        isheet.cell(
            form_index_start,
            1,
            f"     "
            + f"学校名称：{self.bill.profile.org_name}"
            + f"                "
            + f"{time_end.year} 年 {time_end.month} 月 {time_end.day} 日"
            + f"              ",
        )

        for row in isheet.iter_rows(
            min_row=fentry_index_start,
            max_row=fentry_index_end,
            min_col=1,
            max_col=9,
        ):
            for cell in row:
                cell.value = ""
                cell.alignment = self.cell_alignment0
                cell.border = self.cell_border0

        for findex, food in enumerate(foods):
            row_index = fentry_index_start + findex
            if isheet.cell(row_index + 1, 1).value.replace(" ", "") == "合计":
                isheet.insert_rows(row_index + 1, 1)
            isheet.cell(row_index, 1, food.name)
            isheet.cell(row_index, 2, food.get_unit_name())
            isheet.cell(row_index, 3, food.get_remainder())
            isheet.cell(row_index, 4, food.get_remainder() * food.unit_price)
            isheet.cell(row_index, 5, food.get_remainder())
            isheet.cell(row_index, 6, food.get_remainder() * food.unit_price)

        for form_index_n in range(form_indexes_n + 1, len(form_indexes)):
            form_index = form_indexes[form_index_n]
            form_index_start, form_index_end = form_index
            food_index_start = form_index_start + 3
            food_index_end = form_index_end - 1
            for row in isheet.iter_rows(
                min_row=food_index_start,
                max_row=food_index_end,
                min_col=1,
                max_col=9,
            ):
                for cell in row:
                    cell.value = ""

        self.format_inventory_sheet()

        wb = self.get_workbook()
        wb.active = isheet
        print_info(_("Sheet '%s' was updated.") % (self.inventory_sheet_name))

    def update_check_sheet_by_time_node_m1(self):
        cksheet = self.get_check_sheet()
        rfoods = self.food.get_foods_from_pre_consuming_sheet_m1()
        time_node = self.bill.time_node
        time_start, time_end = time_node
        rfoods = [f for f in rfoods if f.get_remainder() > 0.0]

        for food in rfoods:
            food_exists = False
            for row in cksheet.iter_rows(
                min_row=2, max_row=cksheet.max_row + 1, min_col=1, max_col=2
            ):
                if row[0].value == food.fid and row[
                    1
                ].value == time_end.strftime("%Y%m%d"):
                    food_exists = True
                    break
            if food_exists:
                continue
            cksheet.insert_rows(2, 1)
            cksheet.cell(2, 1, food.fid)
            cksheet.cell(2, 2, time_end.strftime("%Y%m%d"))
            cksheet.cell(2, 3, food.name)
            cksheet.cell(2, 4, food.get_remainder())
            cksheet.cell(2, 5, food.get_remainder() * food.unit_price)
            cksheet.cell(2, 6, "Y")

        wb = self.get_workbook()
        wb.active = cksheet
        print_info(_("Sheet '%s' was updated.") % (self.check_sheet_name))

    def unmerge_cells_of_sheet(self, sheet):
        if isinstance(sheet, str):
            sheet = self.get_sheet(sheet)
        merged_ranges = list(sheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            sheet.unmerge_cells(str(cell_group))

    def update_cover_sheet(self):
        time_start, time_end = self.bill.time_node
        cvsheet = self.get_conver_sheet()
        cvsheet.cell(
            1,
            1,
            self.bill.org_name
            + f"{time_end.year}年{time_end.month}月份食堂食品采购统计表",
        )
        foods = self.food.get_food_list_from_check_sheet()
        foods = [
            f
            for f in foods
            if (
                not f.is_residue
                and self.bill.times_are_same_year_month(f.check_date, time_end)
            )
        ]
        wfoods = [f for f in foods if not f.is_negligible]
        uwfoods = [f for f in foods if f.is_negligible]
        total_price = 0.0
        for row in cvsheet.iter_rows(
            min_row=3, max_row=9, min_col=1, max_col=3
        ):
            class_name = row[0].value.replace(" ", "")
            _total_price = 0.0
            for f in foods:
                if f.get_base_class_name() == class_name:
                    _total_price += f.count * f.unit_price
            cvsheet.cell(row[0].row, 2, _total_price)

            total_price += _total_price
        cvsheet.cell(10, 2, total_price)

        w_seasoning_total_price = sum(
            [
                f.count * f.unit_price
                for f in wfoods
                if ("调味" in f.get_base_class_name())
            ]
        )
        unw_seasoning_total_price = sum(
            [
                f.count * f.unit_price
                for f in uwfoods
                if ("调味" in f.get_base_class_name())
            ]
        )

        cvsheet.cell(
            8,
            3,
            f"入库：{w_seasoning_total_price:.2f}元；"
            + f"未入库：{unw_seasoning_total_price:.2f}元",
        )
        if "昌盛" in self.bill.suppliers and "雪兰" in self.bill.suppliers:
            self.update_cover_sheet_for_cangsheng_xuelan(
                cvsheet, foods, wfoods, uwfoods, total_price
            )

        wb = self.get_workbook()
        wb.active = cvsheet

        print_info(_("Sheet '%s' was updated.") % self.cover_sheet_name)

    def get_food_form_index_by_time_node_m1(self, sheet):
        _, time_end = self.bill.time_node
        indexes = self.get_food_form_indexes(sheet)
        _index_range = indexes[time_end.month - 1]
        return _index_range

    def get_food_form_indexes(self, sheet):
        indexes = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
        ):
            if row[0].value and "材料名称" in str(row[0].value).replace(" ", ""):
                indexes.append([row[0].row + 3, None])

            if row[2].value and "合计" in str(row[2].value).replace(" ", ""):
                indexes[-1][1] = row[0].row + 1
        return indexes

    def get_residual_foods_by_month_m1(self):
        time_nodes = self.bill.get_time_nodes()
        time_start, time_end = time_nodes[-1]
        time_end_mm1 = datetime(time_end.year, time_end.month, 1) + timedelta(
            days=-1
        )
        time_nodes_mm1 = [
            t
            for t in time_nodes
            if self.bill.times_are_same_year_month(t[0], time_end_mm1)
        ]

        foods = self.food.get_food_list_from_check_sheet()

        if len(time_nodes_mm1) < 1:
            time_end_mm1 = time_start + timedelta(days=-1)
        else:
            time_nodes_mm1 = sorted(time_nodes_mm1, key=lambda t: t[1])
            time_end_mm1 = time_nodes_mm1[-1][1]

        foods = [
            f for f in foods if (f.is_residue and f.check_date == time_end_mm1)
        ]
        return foods

    def get_food_sheet(self, name):
        sheet = None
        _, time_end = self.bill.time_node
        if self.includes_sheet(name):
            sheet = self.get_sheet(name)
        else:
            wb = self.get_workbook()
            sheet = wb.copy_worksheet(self.get_food_sheet0())
            sheet.title = name
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
            ):
                if row[0].value and "材料名称" in str(row[0].value).replace(
                    " ", ""
                ):
                    row[0].value = (
                        f"材料名称：{name}" + f"（{self.food.get_unit_name(name)}）"
                    )
                    sheet.cell(row[0].row + 1, 1, f"{time_end.year}年")
        return sheet

    def format_food_sheet(self, sheet):
        if isinstance(sheet, str):
            sheet = self.get_food_sheet(sheet)
        self.unmerge_cells_of_sheet(sheet)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=14,
        ):
            sheet.row_dimensions[row[0].row].height = 15.75
            if row[0].value and "入库、出库台账" in str(row[0].value):
                sheet.row_dimensions[row[0].row].height = 27
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=13,
                )

            if row[0].value and "年" in str(row[0].value):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=2,
                )

            if row[3].value and "入库" in str(row[3].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=4,
                    end_column=6,
                )

            if row[6].value and "出库" in str(row[6].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=7,
                    end_column=9,
                )

            if row[9].value and "库存" in str(row[9].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=10,
                    end_column=12,
                )

            if row[12].value and "编号" in str(row[12].value).replace(" ", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=13,
                    end_column=13,
                )

    def update_food_sheets_by_time_nodes_m1(self):
        time_nodes = self.bill.get_time_nodes()
        time_start, time_end = time_nodes[-1]
        cfoods = (
            self.food.get_foods_from_pre_consuming_sheet_by_time_nodes_m1()
        )
        cfood_names = list(set([f.name for f in cfoods]))
        u_month = time_end.month
        days_num = calendar.monthrange(time_end.year, u_month)[1]
        wb = self.get_workbook()

        if len(time_nodes) > 1:
            time_end_m2 = time_nodes[-2][1]
        else:
            time_end_m2 = time_start + timedelta(days=-1)
        rfoods = self.get_residual_foods_by_month_m1()
        rfoods_names = list(set([f.name for f in rfoods]))
        for rfood_name in rfoods_names:
            sheet = self.get_food_sheet(rfood_name)
            form_index_range = self.get_food_form_index_by_time_node_m1(sheet)
            index_start, index_end = form_index_range
            entry_index = 0
            for food in rfoods:
                if food.name == rfood_name:
                    row_index = index_start + entry_index
                    sheet.cell(
                        row_index,
                        3,
                        ("上年结转" if time_end.month == 1 else "上月结转"),
                    )
                    for col_index in [1, 2, 4, 5, 6, 7, 8, 9]:
                        sheet.cell(row_index, col_index, "")
                    sheet.cell(row_index, 10, food.count)
                    sheet.cell(row_index, 11, food.unit_price)
                    sheet.cell(row_index, 12, food.count * food.unit_price)

                    entry_index += 1

        for cfood_name in cfood_names:
            entry_index = 0
            sheet = self.get_food_sheet(cfood_name)
            self.unmerge_cells_of_sheet(sheet)
            index_range = self.get_food_form_day_index_by_time_node_m1(sheet)
            index_start, index_end = index_range
            for day_n in range(1, days_num + 1):
                time_node = datetime(time_end.year, time_end.month, day_n)
                for food in [f for f in cfoods if (f.name == cfood_name)]:
                    _dates = [d for d, c in food.consuming_list]
                    if time_node in _dates:
                        row_index = index_start + entry_index
                        _date, _count = [
                            c for c in food.consuming_list if c[0] == time_node
                        ][0]
                        _remainder = food.count - sum(
                            [
                                c
                                for d, c in food.consuming_list
                                if d <= time_node
                            ]
                        )

                        sheet.cell(row_index, 2, time_node.day)
                        sheet.cell(row_index, 6, "")
                        sheet.cell(row_index, 7, _count)
                        sheet.cell(row_index, 8, food.unit_price)
                        sheet.cell(row_index, 9, _count * food.unit_price)
                        sheet.cell(row_index, 10, _remainder)
                        sheet.cell(row_index, 11, food.unit_price)
                        sheet.cell(row_index, 12, _remainder * food.unit_price)

                        entry_index += 1
                    if food.check_date == time_node:
                        row_index = index_start + entry_index

                        sheet.cell(row_index, 2, time_node.day)
                        sheet.cell(row_index, 4, food.count)
                        sheet.cell(row_index, 5, food.unit_price)
                        sheet.cell(row_index, 6, food.count * food.unit_price)
                        sheet.cell(row_index, 9, "")
                        sheet.cell(row_index, 10, food.count)
                        sheet.cell(row_index, 11, food.unit_price)
                        sheet.cell(row_index, 12, food.count * food.unit_price)

                        entry_index += 1

                    if food.check_date == time_node or time_node in _dates:
                        if "合计" in str(sheet.cell(row_index + 1, 3).value):
                            sheet.insert_rows(row_index + 1, 1)

            self.format_food_sheet(sheet)

            all_food_names = [
                f.name for f in self.food.get_food_list_from_check_sheet()
            ]

            wb = self.get_workbook()
            wb.active = sheet

            print_info(_("Sheet '%s' was updated.") % sheet.title)

        for name in all_food_names:
            if self.includes_sheet(name):
                sheet = self.get_sheet(name)
                sheet.sheet_properties.tabColor = "0" * 8
        for name in cfood_names:
            sheet = self.get_food_sheet(name)
            sheet.sheet_properties.tabColor = secrets.token_hex(4)

    def get_food_form_day_index_by_time_node_m1(self, sheet):
        indexes = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
        ):
            if (
                row[2].value
                and "结转" in str(row[2].value).replace(" ", "")
                and not sheet.cell(row[0].row + 1, 3).value
            ):
                indexes.append([row[0].row + 1, None])

            if row[2].value and "合计" in str(row[2].value).replace(" ", ""):
                indexes[-1][1] = row[0].row - 1
        _, time_end = self.bill.time_node
        index_range = indexes[time_end.month - 1]
        return index_range

    def update_cover_sheet_for_cangsheng_xuelan(
        self, cvsheet, foods, wfoods, uwfoods, total_price
    ):
        egg_milk_total_price = sum(
            [
                f.count * f.unit_price
                for f in foods
                if "蛋奶" in f.get_base_class_name()
            ]
        )
        xl_milk_total_price = sum(
            [f.count * f.unit_price for f in foods if "雪兰" in f.name]
        )

        cvsheet.cell(
            5,
            3,
            f"昌盛：{(egg_milk_total_price - xl_milk_total_price):.2f}元；"
            + f"雪兰：{xl_milk_total_price:.2f}元",
        )

        cvsheet.cell(
            10,
            3,
            f"昌盛：{total_price-xl_milk_total_price:.2f}元；"
            + f"雪兰：{xl_milk_total_price:.2f}元",
        )

    def clean_food_count(self, name, count, unit):
        return (
            count * 50
            if unit == "包" and "小町" in name
            else count * 360
            if unit == "件" and "大号鸡蛋" == name
            else count * 8
            if unit == "件" and "三全奶香馒头" in name
            else count * 4
            if unit == "件" and "滇雪菜家村上等菜籽油" in name
            else count
        )

    def get_changsheng_purchase_properties(self, fpath):
        if not fpath.split(".")[-1] in self.spreadsheet_ext_names:
            return None
        chwb_fpath = fpath
        ef = pd.ExcelFile(chwb_fpath)
        sheet_names = ef.sheet_names
        ef.close()
        for sheet_name in sheet_names:
            if sheet_name in self.purchase_sheet_names:
                chwb0 = load_workbook(chwb_fpath, read_only=True)
                cssheet0 = chwb0[sheet_name]
                cs_dates = []
                row_index = 1
                while True:
                    _date = cssheet0.cell(row_index, 2).value
                    if not _date:
                        break
                    cs_dates.append(str(_date))
                    row_index += 1
                row_index = 1

                cs_dates = sorted(
                    list(
                        set(
                            [
                                datetime.strptime(d, "%Y-%m-%d")
                                for d in cs_dates
                                if re.search(r"\d{4}-\d{2}-\d{2}", d)
                            ]
                        )
                    )
                )
                chwb0.close()
                return (sheet_name, cs_dates)

        return None

    def get_changsheng_sheet_name(self, wb):
        sheet_names = []
        if isinstance(wb, str):
            ef = pd.ExcelFile(wb)
            sheet_names = ef.sheet_names
            ef.close()
        else:
            sheet_names = wb.sheetnames
        for sn in self.purchase_sheet_names:
            if sn in sheet_names:
                return sn
        return None

    def get_changsheng_properties_by_dir(self, fdpath=None):
        fd_path = self.purchase_workbook_fd_path or fdpath
        properties = []
        if not Path(fd_path).is_dir():
            return None

        for _file in os.listdir(fd_path):
            if _file.split(".")[-1] in self.spreadsheet_ext_names:
                chwb_fpath = (Path(fd_path) / _file).as_posix()
                print_info(_("Spreadsheet %s is being tested.") % _file)
                pinfo = self.get_changsheng_purchase_properties(chwb_fpath)
                if not pinfo:
                    continue
                sheet_name, ptimes = pinfo
                if ptimes:
                    print_info(
                        _(
                            "The food purchasing times of preadsheet {0} is {1} ."
                        ).format(
                            _file,
                            " | ".join(
                                [
                                    ptime.strftime("%Y.%m.%d")
                                    for ptime in ptimes
                                ]
                            ),
                        )
                    )
                    properties.append([_file, sheet_name, chwb_fpath, ptimes])

        return properties if properties else None

    def read_changsheng_foods_by_time_node(self, fd_path=None, time_node=None):
        global Food
        fd_path = self.purchase_workbook_fd_path or fd_path
        time_node = time_node or self.bill.time_node
        time_start, time_end = time_node
        seeking_dpath0 = (Path.home() / "Downloads").as_posix()
        if not fd_path:
            print_info(
                _(
                    "Please enter the 'purchase list file path' of "
                    + "spreadsheet Changsheng provided, "
                    + "or enter the directory path and then {app_name} will "
                    + "read all spreadsheets."
                    + " (default: '{seeking_dpath0}')"
                ).format(app_name=app_name, seeking_dpath0=seeking_dpath0)
            )
            fd_path = input(">_ ")

        if fd_path.replace(" ", "") == "":
            fd_path = seeking_dpath0

        if fd_path.startswith("~"):
            fd_path = Path.home().as_posix() + fd_path[1:]

        if not Path(fd_path).exists():
            print_error(_("File or directory '%s' doesn't exist.") % (fd_path))
            return None

        chwb = None
        cssheet = None
        ck_t0, ck_t1 = self.bill.get_check_times_of_time_node()

        if Path(fd_path).is_dir():
            print_info(_("Entered directory: %s") % fd_path)
            csproperties = self.get_changsheng_properties_by_dir(fd_path)
            if not csproperties:
                return None
            for csproperty in csproperties:
                file_name, sheet_name, file_path, purchase_times = csproperty
                for ptime in purchase_times:
                    if ck_t0 <= ptime <= ck_t1:
                        fd_path = file_path
                        cssheet = sheet_name
                        break
                if cssheet:
                    break
            if not cssheet:
                return None

        else:
            cssheet = self.get_changsheng_sheet_name(fd_path)

        chwb = load_workbook(fd_path, read_only=True)
        cssheet = chwb[cssheet]

        food_name_index = 0
        food_count_index = 0
        food_total_price_index = 0
        food_unit_index = 0
        food_check_date_index = 0
        food_neglect_mark_index = 0
        food_residue_mark_index = 0

        col_index = 1
        break_index = 0
        while True:
            cell = cssheet.cell(1, col_index)
            cell_value = cell.value
            col_index = col_index - 1
            if not cell_value:
                if break_index > 3:
                    break
                break_index += 1
            cell_value = str(cell_value.replace(" ", ""))
            if cell_value in ["商品名称"]:
                food_name_index = col_index
                continue

            elif cell_value in ["单位", "订货单位"]:
                food_unit_index = col_index
                continue

            elif cell_value in ["数量", "记账数量"]:
                food_count_index = col_index
                continue

            elif cell_value in ["金额", "折前金额"]:
                food_total_price_index = col_index
                continue

            elif cell_value in ["送货日期"]:
                food_check_date_index = col_index
                continue
            elif cell_value in ["忽略", "不计", "非入库", "可忽略", "非盘点"]:
                food_neglect_mark_index = col_index
                continue
            elif cell_value in ["上季结余", "是剩余", "是结余", "上年结余", "剩余", "结余"]:
                food_residue_mark_index = col_index
                continue

            col_index += 1

        csfoods = []
        is_residue = False

        row_index = 1
        col_index = 1
        for row in cssheet.iter_rows(
            min_row=2,
            max_row=cssheet.max_row,
            min_col=1,
            max_col=cssheet.max_column,
        ):
            if row[food_name_index].value:
                check_date = row[food_check_date_index].value
                check_date = (
                    datetime.strptime(check_date, "%Y-%m-%d")
                    if "-" in check_date
                    else datetime.strptime(check_date, "%Y%d%m")
                )
                if not (ck_t0 <= check_date <= ck_t1):
                    continue

                name = row[food_name_index].value
                count = row[food_count_index].value
                unit = row[food_unit_index].value
                total_price = row[food_total_price_index].value
                is_negligible = (
                    not row[food_neglect_mark_index].value is None
                    if food_neglect_mark_index
                    else False
                )
                is_residue = (
                    not row[food_residue_mark_index] is None
                    if food_residue_mark_index
                    else False
                )

                if not is_residue:
                    count = self.clean_food_count(name, count, unit)

                csfoods.append(
                    Food(
                        self.bill,
                        name=name,
                        check_date=check_date,
                        count=count,
                        is_residue=is_residue,
                        total_price=total_price,
                        is_negligible=is_negligible,
                    )
                )

            if (
                not row[food_name_index].value
                and cssheet.cell(row[0].row + 1, food_name_index + 1).value
            ):
                is_residue = True
        chwb.close()
        return csfoods

    def get_inventory_form_index_of_time_node(self):
        indexes = self.get_inventory_form_indexes()
        tn_index = (
            self.bill.get_time_node_index() + self.inventory_form_index_offset
        )
        indexes_len = len(indexes)
        _index = None
        if indexes_len <= tn_index:
            return None
        _index = indexes[tn_index]
        return _index

    def update_inventory_sheet_of_time_node(
        self, fd_path=None, time_node=None
    ):
        fd_path = self.purchase_workbook_fd_path or fd_path
        time_node = time_node or self.bill.time_node
        time_start, time_end = time_node
        isheet = self.get_inventory_sheet()
        foods = self.food.time_node_residue_foods
        (
            iform_index0,
            iform_index1,
        ) = self.get_inventory_form_index_of_time_node()

        if not foods:
            print_warning(_("There is no residue foods."))
            return isheet

        for food in foods:
            isheet.cell(iform_index0, 1, name)
            isheet.cell(iform_index0, 2, self.food.get_unit_name(name))
            isheet.cell(iform_index0, 3, count)
            isheet.cell(iform_index0, 4, total_price)
            isheet.cell(iform_index0, 5, count)
            isheet.cell(iform_index0, 6, total_price)
            if (
                isheet.cell(iform_index0 + 1, 1).value
                and isheet.cell(iform_index0 + 1, 1).value.replace(" ", "")
                == "合计"
            ):
                isheet.insert_rows(iform_index0 + 1, 1)
                r_total_price += total_price
                iform_index0 += 1
                isheet.cell(isht_form_index_end, 4, r_total_price)
                isheet.cell(isht_form_index_end, 6, r_total_price)

        print_info(_("Sheet '%s' was updated.") % self.inventory_sheet_name)

        wb = self.get_workbook()
        wb.active = isheet
        print_info(_("Sheet '%s' was updated.") % self.check_sheet_name)
        return isheet

    def update_purchase_sum_sheet_by_time_node(self):
        time_start, time_end = self.bill.time_node
        pssheet = self.get_purchase_sum_sheet()
        pssheet.cell(
            2,
            1,
            f"编制单位：{self.bill.org_name}"
            + f"        "
            + f"单位：元"
            + f"         "
            + f"{time_end.year}年{time_end.month}月{time_end.day}日",
        )
        pssheet.cell(
            20,
            1,
            f"编制单位：{self.bill.org_name}"
            + f"        "
            + f"单位：元"
            + f"         "
            + f"{time_end.year}年{time_end.month}月{time_end.day}日",
        )
        foods = self.food.get_food_list_from_check_sheet()
        foods = [
            f
            for f in foods
            if (
                not f.is_residue
                and self.bill.times_are_same_year_month(f.check_date, time_end)
            )
        ]
        wfoods = [f for f in foods if not f.is_negligible]
        uwfoods = [f for f in foods if f.is_negligible]
        total_price = 0.0
        for row in pssheet.iter_rows(
            min_row=4, max_row=10, min_col=1, max_col=3
        ):
            class_name = row[0].value.replace(" ", "")
            _total_price = 0.0
            for food in wfoods:
                if food.get_base_class_name() == class_name:
                    _total_price += food.count * food.unit_price
            pssheet.cell(row[0].row, 2, _total_price)
            total_price += _total_price
        total_price_cn = self.bill.convert_num_to_cnmoney_chars(total_price)
        pssheet.cell(11, 1, f"总金额（大写)：{total_price_cn}    ¥{total_price:.2f}")
        pssheet.cell(12, 1, f"经办人：{self.bill.profile.name}  ")

        total_price = sum([f.count * f.unit_price for f in uwfoods])
        total_price_cn = self.bill.convert_num_to_cnmoney_chars(total_price)
        pssheet.cell(27, 2, total_price)
        pssheet.cell(29, 1, f"总金额（大写)：{total_price_cn}    ¥{total_price:.2f}")

        pssheet.cell(30, 1, f"经办人：{self.bill.profile.name}  ")

        wb = self.get_workbook()
        wb.active = pssheet

        print_info(_("Sheet '%s' was updated.") % self.purchase_sum_sheet_name)

    def update_consuming_sum_sheet(self):
        cssheet = self.get_consuming_sum_sheet()
        time_node = self.bill.time_node
        time_start, time_end = time_node
        foods = self.food.get_foods_from_pre_consuming_sheet_by_time_nodes_m1()

        total_price = 0.0
        for row in cssheet.iter_rows(
            min_row=4, max_row=10, min_col=1, max_col=3
        ):
            class_name = row[0].value.replace(" ", "")
            _total_price = 0.0
            for food in foods:
                if food.get_base_class_name() == class_name:
                    _total_price += sum(
                        [
                            _count * food.unit_price
                            for _date, _count in food.consuming_list
                        ]
                    )
            total_price += _total_price
            cssheet.cell(row[0].row, 2, _total_price)
            cssheet.cell(
                row[0].row, 2
            ).number_format = numbers.FORMAT_NUMBER_00

        total_price_cn = self.bill.convert_num_to_cnmoney_chars(total_price)
        cssheet.cell(
            2,
            1,
            f"编制单位：{self.bill.org_name}       "
            + f"单位：元         "
            + f"{time_end.year}年{time_end.month}月{time_end.day}日",
        )
        cssheet.cell(
            11, 1, (f"总金额（大写)：{total_price_cn}    " + f"¥{total_price:.2f}")
        )
        cssheet.cell(12, 1, f"经办人：{self.bill.profile.name}  ")

        wb = self.get_workbook()
        wb.active = cssheet

        print_info(
            _("Sheet '%s' was updated.") % self.consuming_sum_sheet_name
        )

    def update_consuming_sheet_by_time_node_m1(self, quiet=False):
        self.update_pre_consuming_sheet_m1(quiet)
        csheet = self.get_consuming_sheet()
        form_indexes = self.get_consuming_form_indexes()
        time_start, time_end = self.bill.time_node
        foods = self.food.get_foods_from_pre_consuming_sheet_m1()
        class_names = self.get_base_class_names()

        merged_ranges = list(csheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            csheet.unmerge_cells(str(cell_group))

        max_time_index = 0
        for time_index in range(0, (time_end - time_start).days + 1):
            max_time_index = time_index + 1
            time_node = time_start + timedelta(days=time_index)
            form_index = form_indexes[time_index]
            form_index_start, form_index_end = form_index
            food_index_start = form_index_start + 2
            food_index_end = form_index_end - 1
            food_index_len = food_index_end - food_index_start + 1
            tfoods = [
                food
                for food in foods
                if time_node
                in [_date for _date, _count in food.consuming_list]
            ]
            tfoods_classes = [f.get_base_class_name() for f in tfoods]

            classes_without_food = [
                _name for _name in class_names if not _name in tfoods_classes
            ]

            tfoods_len = len(tfoods)
            consuming_n = self.get_consuming_n_by_time_node(time_node)
            csheet.cell(
                form_index_start,
                4,
                f"{time_node.year}年 {time_node.month} 月 {time_node.day} 日  "
                + f"单位：元",
            )
            csheet.cell(
                form_index_start,
                7,
                f"编号：C{time_node.month:0>2}{consuming_n:0>2}",
            )

            row_difference = (
                tfoods_len + len(classes_without_food) - food_index_len
            )

            if row_difference > 0:
                csheet.insert_rows(food_index_start + 1, row_difference)
                form_indexes = self.get_consuming_form_indexes()
                form_index_end += row_difference
                food_index_end += row_difference
                row_difference = 0

            for row in csheet.iter_rows(
                min_row=food_index_start,
                max_row=food_index_end,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

            fentry_index = food_index_start

            for class_name in class_names:
                class_foods = [
                    food
                    for food in tfoods
                    if (food.get_base_class_name() == class_name)
                ]

                fentry_index_start = fentry_index
                if len(class_foods) < 1:
                    csheet.cell(fentry_index_start, 1, class_name)
                    fentry_index = fentry_index_start + 1
                    continue

                class_consuming_count = 0.0
                for food in class_foods:
                    for _date, _count in food.consuming_list:
                        if _date == time_node:
                            class_consuming_count += _count * food.unit_price

                class_foods_len = len(class_foods)
                if class_name == class_names[0] and row_difference < 0:
                    class_foods_len += abs(row_difference) - len(
                        classes_without_food
                    )
                fentry_index_end = fentry_index_start - 1 + class_foods_len

                csheet.cell(fentry_index_start, 1, class_name)
                csheet.cell(fentry_index_start, 7, class_consuming_count)
                csheet.cell(
                    fentry_index_start, 7
                ).number_format = numbers.FORMAT_NUMBER_00

                for findex, food in enumerate(class_foods):
                    consuming_count = [
                        _count
                        for _date, _count in food.consuming_list
                        if _date == time_node
                    ][0]
                    frow_index = fentry_index_start + findex
                    csheet.cell(frow_index, 2, food.name)
                    csheet.cell(frow_index, 3, food.get_unit_name())
                    csheet.cell(frow_index, 4, consuming_count)
                    csheet.cell(frow_index, 5, food.unit_price)
                    csheet.cell(
                        frow_index, 6, consuming_count * food.unit_price
                    )
                    csheet.cell(
                        frow_index, 4
                    ).number_format = numbers.FORMAT_NUMBER_00
                    csheet.cell(
                        frow_index, 5
                    ).number_format = numbers.FORMAT_NUMBER_00
                    csheet.cell(
                        frow_index, 6
                    ).number_format = numbers.FORMAT_NUMBER_00

                fentry_index = fentry_index_end + 1

            tfoods_total_price = 0.0
            for food in tfoods:
                for _date, _count in food.consuming_list:
                    if _date == time_node:
                        tfoods_total_price += _count * food.unit_price
            csheet.cell(form_index_end, 6, tfoods_total_price)
            csheet.cell(form_index_end, 7, tfoods_total_price)

        if len(form_indexes) > max_time_index:
            for time_index in range(max_time_index, len(form_indexes)):
                form_index_start, form_index_end = form_indexes[time_index]
                food_index_start, food_index_end = (
                    form_index_start + 2,
                    form_index_end - 1,
                )
                for row in csheet.iter_rows(
                    min_row=food_index_start,
                    max_row=food_index_end,
                    min_col=2,
                    max_col=7,
                ):
                    for cell in row:
                        cell.value = ""

        self.format_consuming_sheet()

        wb = self.get_workbook()
        wb.active = csheet
        print_info(_("Sheet '%s' was updated.") % self.consuming_sheet_name)

    def update_pre_consuming_sheet(self, quiet=False):
        sheet = self.get_pre_consuming_sheet_m1()
        sheet_title = sheet.title
        foods = self.food.get_non_negligible_foods_by_time_node_m1()
        time_start, time_end = self.bill.time_node
        wb = self.get_workbook()
        row_index_offset = 3
        col_index_offset = 6
        rc_index = 0
        days_difference = (time_end - time_start).days

        for day in range(0, days_difference + 1):
            time_header = (time_start + timedelta(days=day)).strftime(
                "%Y.%m.%d"
            )
            cell = sheet.cell(1, rc_index + col_index_offset)
            cell.value = time_header
            cell.number_format = numbers.FORMAT_TEXT
            rc_index += 1

        rc_index = 0
        for row in sheet.iter_rows(
            max_col=5,
            min_row=row_index_offset,
            max_row=row_index_offset + len(foods),
        ):
            if rc_index > len(foods) - 1:
                break
            food = foods[rc_index]
            row[0].value = food.fid
            row[1].value = food.get_name_withresidue_mark()
            row[2].value = food.count
            row[4].value = food.unit_price
            rc_index += 1

        wb.active = sheet

        if not quiet:
            self.save_workbook()
            print_info(
                f"Sheet '{sheet_title}' was updated.\n"
                + f"Press any key to continue when you have "
                + f"completed the foods allocation."
            )
            input()

        self.clear_workbook()
        sheet = self.get_sheet(sheet_title)
        return sheet

    def get_pre_consuming_sheet_m1(self):
        return self.get_pre_consuming_sheet_by_time_node_m1()

    def get_pre_consuming_sheet_by_time_node_m1(self):
        return self.get_pre_consuming_sheet_by_time_node(
            self.bill.get_time_nodes()[-1]
        )

    def get_pre_consuming_sheet_name_by_time_node(self, time_node):
        time_start, time_end = time_node
        sheet_name = (
            self.pre_consuming_sheet_name_prefix + time_start.strftime("%m%d")
        )
        wb = self.get_workbook()

        for _name in self.get_sheet_names():
            if _name.startswith(sheet_name):
                return _name

        sheet_name += time_end.strftime("%m%d")
        return sheet_name

    def get_pre_consuming_sheet_by_time_node(self, time_node):
        time_start, time_end = time_node
        sheet_name = (
            self.pre_consuming_sheet_name_prefix + time_start.strftime("%m%d")
        )
        wb = self.get_workbook()

        for _name in self.get_sheet_names():
            if _name.startswith(sheet_name):
                return self.get_sheet(_name)

        sheet_name += time_end.strftime("%m%d")
        pre_consuming_sheet = wb.copy_worksheet(
            self.get_pre_consuming_sheet0()
        )
        pre_consuming_sheet.title = sheet_name
        color_fill = PatternFill(
            start_color="ff00ea", end_color="ff00ea", fill_type="solid"
        )
        pre_consuming_sheet.conditional_formatting.add(
            "D3:D1024",
            CellIsRule(
                operator="lessThanOrEqual",
                formula=["0"],
                stopIfTrue=True,
                fill=color_fill,
            ),
        )
        pre_consuming_sheet.sheet_view.zoom = 80
        return pre_consuming_sheet

    def set_warehousing_form_index_offset(self, offset=0):
        self.warehousing_form_index_offset = offset

    def set_inventory_form_index_offset(self, offset=0):
        self.inventory_form_index_offset = offset

    @property
    def food_list(self):
        return self.bill.get_food_list()

    def get_warehousing_sheet(self):
        return self.get_sheet(self.warehousing_sheet_name)

    def get_unwarehousing_sheet(self):
        return self.get_sheet(self.unwarehousing_sheet_name)

    def get_consuming_sheet(self):
        return self.get_sheet(self.consuming_sheet_name)

    def get_inventory_sheet(self):
        return self.get_sheet(self.inventory_sheet_name)

    def get_food_sheet0(self):
        return self.get_sheet(self.food_sheet0_name)

    def get_consuming_form_indexes(self):
        csheet = self.get_consuming_sheet()
        indexes = []
        row_index = 1
        for row in csheet.iter_rows(max_row=csheet.max_row + 1, max_col=9):
            if row[0].value:
                if row[0].value.replace(" ", "") == "出库单":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def get_unwarehousing_form_indexes(self):
        unwsheet = self.get_unwarehousing_sheet()
        indexes = []
        row_index = 1
        for row in unwsheet.iter_rows(max_row=unwsheet.max_row + 1, max_col=7):
            if row[0].value and "未入库明细表" in row[0].value.replace(" ", ""):
                indexes.append([row_index + 1, 0])

            if row[1].value and "合计" in row[1].value.replace(" ", ""):
                indexes[-1][1] = row_index

            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def get_inventory_form_indexes(self):
        isheet = self.get_inventory_sheet()
        indexes = []
        row_index = 1
        for row in isheet.iter_rows(max_row=isheet.max_row + 1, max_col=8):
            if row[0].value:
                if row[0].value.replace(" ", "") == "食材盘存表":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def get_warehousing_form_indexes(self):
        wsheet = self.get_warehousing_sheet()
        indexes = []
        row_index = 1
        for row in wsheet.iter_rows(max_row=wsheet.max_row + 1, max_col=8):
            if row[0].value:
                if row[0].value.replace(" ", "") == "入库单":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def format_consuming_sheet(self):
        csheet = self.get_consuming_sheet()
        merged_ranges = list(csheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            csheet.unmerge_cells(str(cell_group))

        for row in csheet.iter_rows(
            min_row=1, max_row=csheet.max_row, min_col=1, max_col=8
        ):
            if row[0].value and row[0].value.replace(" ", "") == "出库单":
                csheet.row_dimensions[row[0].row].height = 21
                csheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )
                csheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=4,
                    end_column=6,
                )
                csheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=7,
                    end_column=8,
                )

            if row[0].value and row[0].value.replace(" ", "").endswith("类"):
                row[6].number_format = numbers.FORMAT_NUMBER_00
                for _row in csheet.iter_rows(
                    min_row=row[0].row + 1,
                    max_row=csheet.max_row + 1,
                    min_col=1,
                    max_col=1,
                ):
                    if _row[0].value and (
                        _row[0].value.replace(" ", "").endswith("类")
                        or _row[0].value.replace(" ", "") == "合计"
                    ):
                        csheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=1,
                            end_column=1,
                        )
                        csheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=7,
                            end_column=7,
                        )
                        break

            if row[0].value and "审核人" in row[0].value.replace(" ", ""):
                csheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )

        wb = self.get_workbook()
        wb.active = csheet

        print_info(_("Sheet '%s' was formatted.") % self.consuming_sheet_name)

    def format_warehousing_sheet(self):
        wsheet = self.get_warehousing_sheet()
        merged_ranges = list(wsheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            wsheet.unmerge_cells(str(cell_group))

        for row in wsheet.iter_rows(
            min_row=1, max_row=wsheet.max_row, min_col=1, max_col=8
        ):
            if row[0].value and row[0].value.replace(" ", "") == "入库单":
                wsheet.row_dimensions[row[0].row].height = 21
                wsheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )
                wsheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=4,
                    end_column=6,
                )
                wsheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=7,
                    end_column=8,
                )

            if row[0].value and row[0].value.replace(" ", "").endswith("类"):
                row[6].number_format = numbers.FORMAT_NUMBER_00
                for _row in wsheet.iter_rows(
                    min_row=row[0].row + 1,
                    max_row=wsheet.max_row + 1,
                    min_col=1,
                    max_col=1,
                ):
                    if _row[0].value and (
                        _row[0].value.replace(" ", "").endswith("类")
                        or _row[0].value.replace(" ", "") == "合计"
                    ):
                        wsheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=1,
                            end_column=1,
                        )
                        wsheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=7,
                            end_column=7,
                        )
                        break

            if row[0].value and "审核人" in row[0].value.replace(" ", ""):
                wsheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )

        wb = self.get_workbook()
        wb.active = wsheet

        print_info(
            _("Sheet '%s' was formatted.") % self.warehousing_sheet_name
        )

    def update_unwarehousing_sheet_by_time_node_m1(self):
        unwsheet = self.get_unwarehousing_sheet()
        form_indexes = self.get_unwarehousing_form_indexes()
        foods = self.food.get_food_list_from_check_sheet()
        time_nodes = self.bill.get_time_nodes()
        time_node = time_nodes[-1]
        time_start, time_end = time_node
        foods = [
            f
            for f in foods
            if (
                not f.is_residue
                and f.is_negligible
                and self.bill.times_are_same_year_month(
                    f.check_date, time_start
                )
            )
        ]
        foods = sorted(foods, key=lambda f: f.check_date)
        row_indexes = []
        for form_index in form_indexes:
            form_index_start, form_index_end = form_index
            unwsheet.cell(form_index_start, 1, f" 学校名称：{self.bill.org_name}")
            unwsheet.cell(
                form_index_start,
                4,
                f"        "
                + f"{time_end.year} 年 {time_end.month} 月 "
                + f"{time_end.day} 日"
                + f"               ",
            )
            row_index_start = form_index_start + 2
            row_index_end = form_index_end - 1
            row_indexes += list(range(row_index_start, row_index_end + 1))

        for row_index in row_indexes:
            for col_index in range(1, 7 + 1):
                unwsheet.cell(row_index, col_index, "")

        total_price = 0.0
        use_forms = False

        for _index, row_index in enumerate(row_indexes):
            food = foods[_index]
            total_price += food.total_price
            unwsheet.cell(row_index, 1, food.check_date.strftime("%Y.%m.%d"))
            unwsheet.cell(row_index, 2, food.name)
            unwsheet.cell(row_index, 3, food.get_unit_name())
            unwsheet.cell(row_index, 4, food.count)
            unwsheet.cell(row_index, 5, food.unit_price)
            unwsheet.cell(row_index, 6, food.total_price)
            unwsheet.cell(
                row_index, 5
            ).number_format = numbers.FORMAT_NUMBER_00
            unwsheet.cell(
                row_index, 6
            ).number_format = numbers.FORMAT_NUMBER_00
            if (
                str(unwsheet.cell(row_index + 1, 2).value)
                .replace(" ", "")
                .endswith("合计")
                and len(foods) - 1 > _index
            ):
                unwsheet.cell(row_index + 1, 2, "合计")
                unwsheet.cell(row_index + 1, 6, total_price)
                use_forms = True

            if len(foods) - 1 == _index:
                for row in unwsheet.iter_rows(
                    min_row=row_index,
                    max_row=unwsheet.max_row,
                    min_col=1,
                    max_col=7,
                ):
                    if row[2].value and str(row[2].value).replace(
                        " ", ""
                    ).endswith("合计"):
                        row[2].value = "总合计" if use_forms else "合计"
                        row[6].value = total_price
                        break
                break

        print_info(
            _("Sheet '%s' was updated.") % self.unwarehousing_sheet_name
        )

    def update_warehousing_sheet_by_time_node(self):
        wsheet = self.get_warehousing_sheet()
        foods = self.food.time_node_foods
        form_indexes = self.get_warehousing_form_indexes()
        class_names = self.get_base_class_names()
        time_node = self.bill.time_node

        self.unmerge_cells_of_sheet(wsheet)

        for form_index_start, form_index_end in form_indexes:
            food_index_start = form_index_start + 2
            food_index_end = form_index_end - 1
            for row in wsheet.iter_rows(
                min_row=food_index_start,
                max_row=food_index_end,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""

        foods = [
            f for f in foods if (not f.is_residue and not f.is_negligible)
        ]

        w_times = sorted(
            list(
                set(
                    [
                        food.check_date
                        for food in foods
                        if self.bill.times_are_same_year_month(
                            food.check_date, time_node[1]
                        )
                    ]
                )
            )
        )

        max_time_index = 0
        for windex, w_time in enumerate(w_times):
            time_point = w_time
            max_time_index = windex + 1
            form_index_start, form_index_end = form_indexes[windex]
            food_index_start = form_index_start + 2
            food_index_end = form_index_end - 1
            entry_index = food_index_start
            warehousing_n = windex + 1

            wfoods = [f for f in foods if (f.check_date == time_point)]
            foods_class_names = [f.get_base_class_name() for f in wfoods]
            class_names_without_food = [
                _name
                for _name in class_names
                if not _name in foods_class_names
            ]
            row_difference = (
                len(wfoods)
                + len(class_names_without_food)
                - (food_index_end - food_index_start + 1)
            )

            if row_difference > 0:
                wsheet.insert_rows(food_index_start + 1, row_difference)
                for row in wsheet.iter_rows(
                    min_row=food_index_start + 1,
                    max_row=food_index_start + 1 + row_difference,
                    min_col=1,
                    max_col=8,
                ):
                    for cell in row:
                        cell.alignment = self.cell_alignment0
                        self.border = self.cell_border0

                form_indexes = self.get_warehousing_form_indexes()
                form_index_end += row_difference
                food_index_end = form_index_end - 1
                row_difference = 0

            for row in wsheet.iter_rows(
                min_row=food_index_start,
                max_row=food_index_end,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

            wsheet.cell(form_index_start, 2, self.bill.org_name)
            wsheet.cell(
                form_index_start,
                4,
                f"{time_point.year}年 {time_point.month} 月 "
                + f"{time_point.day} 日  单位：元",
            )
            wsheet.cell(
                form_index_start,
                7,
                f"编号：R{time_point.month:0>2}{warehousing_n:0>2}",
            )

            for class_name in class_names:
                cfoods = [
                    f for f in wfoods if f.get_base_class_name() == class_name
                ]
                cfoods_total_price = sum([f.total_price for f in cfoods])

                wsheet.cell(entry_index, 1, class_name)
                wsheet.cell(entry_index, 7, cfoods_total_price)

                if len(cfoods) < 1:
                    entry_index += 1
                    continue

                for cfindex, cfood in enumerate(cfoods):
                    cfood_row_index = entry_index + cfindex
                    wsheet.cell(cfood_row_index, 2, cfood.name)
                    wsheet.cell(cfood_row_index, 3, cfood.get_unit_name())
                    wsheet.cell(cfood_row_index, 4, cfood.count)
                    wsheet.cell(cfood_row_index, 5, cfood.unit_price)
                    wsheet.cell(cfood_row_index, 6, cfood.total_price)
                    wsheet.cell(
                        cfood_row_index, 5
                    ).number_format = numbers.FORMAT_NUMBER_00
                    wsheet.cell(
                        cfood_row_index, 6
                    ).number_format = numbers.FORMAT_NUMBER_00

                entry_index_end = entry_index + len(cfoods) - 1

                if class_name == class_names[0] and row_difference < 0:
                    entry_index_end = (
                        entry_index_end
                        + abs(row_difference)
                        - len(class_names_without_food)
                    )

                entry_index = entry_index_end + 1
            foods_total_price = sum([f.total_price for f in wfoods])
            wsheet.cell(form_index_end, 6, foods_total_price)
            wsheet.cell(form_index_end, 7, foods_total_price)

        if len(form_indexes) > max_time_index:
            for time_index in range(max_time_index, len(form_indexes)):
                form_index_start, form_index_end = form_indexes[time_index]
                food_index_start, food_index_end = (
                    form_index_start + 2,
                    form_index_end - 1,
                )

                for row in wsheet.iter_rows(
                    min_row=food_index_start,
                    max_row=food_index_end,
                    min_col=2,
                    max_col=7,
                ):
                    for cell in row:
                        cell.value = ""

        self.format_warehousing_sheet()
        wb = self.get_workbook()
        wb.active = wsheet

        print_info(
            _("Sheet '%s' was updated.") % (self.warehousing_sheet_name)
        )

    def get_unit_df(self):
        if not self._unit_df is None:
            return self._unit_df
        self._unit_df = pd.read_excel(
            self.get_main_spreadsheet_path(),
            sheet_name=self.unit_sheet_name,
            names=["Name", "Unit"],
        )
        return self._unit_df

    def get_check_df_from_spreadsheet(self):
        check_df = pd.DataFrame(self.get_check_sheet().values)
        check_df.columns = check_df.iloc[0]
        check_df = check_df[1:]
        check_df = check_df.dropna(axis=0, how="all")
        if check_df.shape[0] == 0:
            return None

        return check_df

    def get_check_df(self):
        if not self._check_df is None:
            return self._check_df
        self._check_df = self.get_check_df_from_spreadsheet()
        return self._check_df

    def clear_check_df(self):
        self._check_df = None

    def get_unit_sheet(self):
        return self.get_sheet(self.unit_sheet_name)

    def get_entry_row_index_of_unit_sheet(self):
        unit_sheet = self.get_unit_sheet()
        row_index = 1
        for row in unit_sheet.iter_rows(
            min_row=1, max_row=unit_sheet.max_row + 1, max_col=1
        ):
            cell = row[0]
            if cell.value == None:
                return row_index
            row_index += 1

        return None

    def add_food_names_to_unit_sheet(self, names):
        unit_sheet = self.get_unit_sheet()
        entry_row_index = self.get_entry_row_index_of_unit_sheet()
        for index, name in enumerate(names):
            unit_sheet.cell(entry_row_index + index, 1, name)

    def get_main_spreadsheet_path(self):
        return (
            self._main_spreadsheet_path
            or self.get_main_spreadsheet_path_of_profile()
        )

    def get_main_spreadsheet_path_of_profile(self):
        s_fpath = workbook0_fpath
        ps_fpath = user_data_dir / self.profile.label / "workbook.xlsx"
        if not ps_fpath.parent.exists():
            os.makedirs(ps_fpath.parent)
            print_info(_("Directory '%s' has been made.") % (ps_fpath.parent))

        if not ps_fpath.exists():
            shutil.copy(s_fpath, ps_fpath)
            print_info(
                _("Workbook '{0}' was copied to '{1}'.").format(
                    s_fpath, ps_fpath
                )
            )
        print_info(_("Workbook '%s' has been used.") % ps_fpath)
        return ps_fpath

    def get_main_spreadsheet0_path(self):
        _path = workbook0_fpath
        return _path

    def set_main_spreadsheet_path(self, file_path=None):
        if not Path(file_path).exists():
            shutil.copy(self.get_main_spreadsheet0_path(), file_path)
        self._main_spreadsheet_path = file_path

    def get_workbook(self):
        if self._workbook:
            return self._workbook
        self._workbook = load_workbook(self.get_main_spreadsheet_path())
        return self._workbook

    def clear_workbook(self):
        self._workbook = None

    def save_workbook(self, info="Saving workbook. . ."):
        print_info(info)
        wb = self.get_workbook()
        wb.save(self.get_main_spreadsheet_path())
        print_info(
            _("Workbook '%s' was saved.") % self.get_main_spreadsheet_path()
        )

    def print_dir_was_created_info(self, dir_path):
        print_info(_("Directory %s was created.") % (dir_path))

    def get_profile_data_dpath(self):
        dpath = user_data_dir / self.bill.profile.label
        if not dpath.exists():
            os.makedirs(dpath)
            self.print_dir_was_created_info(dpath)
        return dpath

    def get_profile_copy_data_dpath(self):
        dpath = self.get_profile_data_dpath()
        dpath = dpath / "copy"
        if not dpath.exists():
            os.makedirs(dpath)
            self.print_dir_was_created_info(dpath)
        return dpath

    def copy_workbook(self, file_path=None):
        if file_path:
            file_path = file_path
        else:
            file_path = self.get_main_spreadsheet_path()
            file_path = file_path.as_posix()
            file_path = file_path.split(os.sep)[-1]
            file_path = file_path[:-5] + f".{uuid.uuid4()}.xlsx"
            file_path = self.get_profile_copy_data_dpath() / file_path

        wb = self.get_workbook()
        wb.save(file_path)
        print_info(_("Workbook '%s' was saved.") % file_path)
        return file_path

    def get_sheet(self, name):
        wb = self.get_workbook()
        return wb[name]

    def get_check_sheet(self):
        return self.get_sheet(self.check_sheet_name)


# The end.
